"""Excel/CSV 파싱 유틸리티.

bytes 입력(FastAPI UploadFile)과 Path 입력 모두 지원합니다.
- Excel: openpyxl 직접 사용 (read_only 모드)
- CSV: chardet 인코딩 감지 + csv.Sniffer 구분자 감지
"""

import csv
import io
from pathlib import Path
from typing import Any

import chardet
import openpyxl
import pandas as pd


def read_to_dataframe(
    content: bytes,
    filename: str,
    *,
    header_row: int = 1,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """바이트 데이터를 DataFrame으로 변환.

    Args:
        content: 파일 바이트 데이터
        filename: 파일명 (확장자로 형식 판별)
        header_row: 헤더 행 번호 (1부터 시작)
        sheet_name: Excel 시트 이름 (None이면 첫 번째 시트)
    """
    headers, rows = extract_headers_and_rows(
        content, filename,
        header_row=header_row,
        sheet_name=sheet_name,
    )
    if not headers:
        return pd.DataFrame()
    return pd.DataFrame(rows, columns=headers)


def extract_headers_and_rows(
    content: bytes,
    filename: str,
    *,
    header_row: int = 1,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """바이트 데이터에서 헤더와 데이터 행 추출.

    Args:
        content: 파일 바이트 데이터
        filename: 파일명 (확장자로 형식 판별)
        header_row: 헤더 행 번호 (1부터 시작)
        sheet_name: Excel 시트 이름 (None이면 첫 번째 시트)
        max_rows: 최대 데이터 행 수 (None이면 전체)

    Returns:
        (headers, rows) 튜플
    """
    suffix = Path(filename).suffix.lower()

    if suffix in (".xlsx", ".xls") or _is_excel_magic(content):
        raw_rows = _read_excel(content, header_row, sheet_name)
    elif suffix == ".csv":
        raw_rows = _read_csv(content, header_row)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {suffix}")

    if not raw_rows:
        return [], []

    raw_headers = raw_rows[0]
    headers = [str(h).strip() for h in raw_headers if h is not None and str(h).strip()]

    data_rows = raw_rows[1:]
    if max_rows is not None:
        data_rows = data_rows[:max_rows]

    rows: list[dict[str, Any]] = []
    for raw_row in data_rows:
        row_dict: dict[str, Any] = {}
        for i, header in enumerate(headers):
            value = raw_row[i] if i < len(raw_row) else None
            row_dict[header] = _clean_value(value)

        if any(v is not None for v in row_dict.values()):
            rows.append(row_dict)

    return headers, rows


def get_sheet_names(content: bytes, filename: str) -> list[str]:
    """Excel 파일의 시트 이름 목록 반환. CSV는 빈 리스트."""
    suffix = Path(filename).suffix.lower()
    if suffix in (".xlsx", ".xls") or _is_excel_magic(content):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    return []


# === 내부 함수 ===

def _is_excel_magic(content: bytes) -> bool:
    """매직 바이트로 Excel 파일 여부 판별."""
    is_xlsx = content[:4] == b"PK\x03\x04"
    is_xls = content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    return is_xlsx or is_xls


def _read_excel(
    content: bytes,
    header_row: int,
    sheet_name: str | None,
) -> list[list[Any]]:
    """Excel 바이트 데이터 읽기."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없음: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows: list[list[Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx >= header_row:
            if any(cell is not None for cell in row):
                rows.append(list(row))

    wb.close()
    return rows


def _read_csv(content: bytes, header_row: int) -> list[list[Any]]:
    """CSV 바이트 데이터 읽기 (인코딩 + 구분자 자동 감지)."""
    detected = chardet.detect(content)
    encoding = detected.get("encoding") or "utf-8"

    try:
        text = content.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        for fallback in ("utf-8-sig", "cp949", "euc-kr", "latin-1"):
            try:
                text = content.decode(fallback)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("파일 인코딩을 인식할 수 없습니다.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    rows: list[list[Any]] = []
    reader = csv.reader(io.StringIO(text), dialect)
    for row_idx, row in enumerate(reader, start=1):
        if row_idx >= header_row:
            if any(cell.strip() for cell in row):
                rows.append(row)

    return rows


_NULL_MARKERS = {"~", "-", "N/A", "n/a", "NA", "없음", "해당없음"}


def _clean_value(value: Any) -> Any:
    """셀 값 정리 (공백 제거, null byte 제거, 특수 NULL 값 처리)."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().replace("\x00", "")
        if not value or value in _NULL_MARKERS:
            return None
        return value
    return value
