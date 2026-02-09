"""Excel/CSV 파싱 유틸리티."""

import csv
from pathlib import Path
from typing import Any

import chardet
import openpyxl


def extract_headers_and_rows(
    file_path: Path,
    header_row: int = 1,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Excel/CSV 파일에서 헤더와 데이터 행 추출.

    Args:
        file_path: 파일 경로
        header_row: 헤더 행 번호 (1부터 시작)
        sheet_name: Excel 시트 이름 (None이면 첫 번째 시트)
        max_rows: 최대 데이터 행 수 (None이면 전체)

    Returns:
        (headers, rows) 튜플
        - headers: 헤더 문자열 리스트
        - rows: dict 리스트 (헤더 -> 값)
    """
    suffix = file_path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        raw_rows = _read_excel(file_path, header_row, sheet_name)
    elif suffix == ".csv":
        raw_rows = _read_csv(file_path, header_row)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {suffix}")

    if not raw_rows:
        return [], []

    # 헤더 추출
    raw_headers = raw_rows[0]
    headers = [str(h).strip() for h in raw_headers if h is not None and str(h).strip()]

    # 데이터 행 추출
    data_rows = raw_rows[1:]
    if max_rows is not None:
        data_rows = data_rows[:max_rows]

    rows: list[dict[str, Any]] = []
    for raw_row in data_rows:
        row_dict: dict[str, Any] = {}
        for i, header in enumerate(headers):
            value = raw_row[i] if i < len(raw_row) else None
            row_dict[header] = _clean_value(value)

        # 빈 행 건너뛰기
        if any(v is not None for v in row_dict.values()):
            rows.append(row_dict)

    return headers, rows


def _read_excel(
    file_path: Path,
    header_row: int,
    sheet_name: str | None,
) -> list[list[Any]]:
    """Excel 파일 읽기."""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없음: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows: list[list[Any]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx >= header_row:
            if any(cell is not None for cell in row):
                rows.append(list(row))

    wb.close()
    return rows


def _read_csv(file_path: Path, header_row: int) -> list[list[Any]]:
    """CSV 파일 읽기."""
    with open(file_path, "rb") as f:
        raw_data = f.read()
        detected = chardet.detect(raw_data)
        encoding = detected.get("encoding", "utf-8")

    rows: list[list[Any]] = []
    with open(file_path, encoding=encoding, newline="") as f:
        sample = f.read(8192)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect)
        for row_idx, row in enumerate(reader, start=1):
            if row_idx >= header_row:
                if any(cell.strip() for cell in row):
                    rows.append(row)

    return rows


def _clean_value(value: Any) -> Any:
    """값 정리."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().replace("\x00", "")
        return value if value else None
    return value
